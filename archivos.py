#Elaborado por: Aarón Ortiz, Alana Calvo.
#Fecha de creación: 14/05/2022 20:25
#Fecha de modificación: 14/05/2022 20:25
#version: 3.10.0
#Importaciones.
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from tkinter import messagebox
def mostrarBaseDatos(participantes):
    """
    Funcionamiento: Crea un reporte con toda la información de los participantes registrados en la base datos.
    entrada(dicc): Base de datos con la información de los participantes.
    Salida: Crea o sobreescribe un archivo excel con toda la información de los participantes.
    """
    try:
        archivoBd = Workbook()
        hoja = archivoBd.active
        hoja.merge_cells("A1:AD1")
        hoja["A1"].alignment = Alignment(horizontal='center')
        hoja["A1"] = "Información de los participantes."
        hoja["A1"].font= Font("Helvetica", size=18, bold=True)
        hoja.merge_cells("A2:C2")
        hoja["A2"].alignment = Alignment(horizontal='center')
        hoja["A2"]= "Fecha nacimiento"
        hoja["A2"].font= Font("Helvetica", size=10, bold=True)
        hoja.merge_cells("D2:F2")
        hoja["D2"]= "Codigo"
        hoja["D2"].font= Font("Helvetica", size=10, bold=True)
        hoja["D2"].alignment= Alignment(horizontal='center')
        hoja.merge_cells("G2:I2")
        hoja["G2"]="Tipo participante"
        hoja["G2"].font= Font("Helvetica", size=10, bold=True)
        hoja.merge_cells("J2:L2")
        hoja["J2"]="Nombre Completo."
        hoja["J2"].font= Font("Helvetica", size=10, bold=True)
        hoja.merge_cells("M2:O2")
        hoja["M2"]="Hobies"
        hoja["M2"].alignment = Alignment(horizontal='center')
        hoja["M2"].font= Font("Helvetica", size=10, bold=True)
        hoja["M2"].alignment= Alignment(horizontal='center')
        hoja.merge_cells("P2:R2")
        hoja["P2"]="Profesión u oficio"
        hoja["P2"].font= Font("Helvetica", size=10, bold=True)
        hoja.merge_cells("S2:U2")
        hoja["S2"]= "Correo electronico"
        hoja["S2"].font= Font("Helvetica", size=10, bold=True)
        hoja.merge_cells("V2:X2")
        hoja["V2"]="Pais"
        hoja["V2"].font= Font("Helvetica", size=10, bold=True)
        hoja.merge_cells("Y2:AA2")
        hoja["Y2"]="Estado"
        hoja["Y2"].font= Font("Helvetica", size=10, bold=True)
        hoja.merge_cells("AB2:AD2")
        hoja["AB2"]="Descripción"
        hoja["AB2"].font= Font("Helvetica", size=10, bold=True)
        fila =3
        for codigo, info in participantes.items():
            hoja.merge_cells(f"A{fila}:C{fila}")
            hoja[f"A{fila}"]=str(info[0])
            hoja[f"A{fila}"].font = Font("Helvetica", size=10)
            hoja.merge_cells(f"D{fila}:F{fila}")
            hoja[f"D{fila}"]= codigo
            hoja[f"D{fila}"].font=Font("Helvetica", size=10)
            hoja.merge_cells(f"G{fila}:I{fila}")
            if info[1]==True:
                hoja[f"G{fila}"]="Adulto Mayor"
            else:
                hoja[f"G{fila}"] = "Voluntario"
            hoja[f"G{fila}"].font= Font("Helvetica", size=10)
            hoja.merge_cells(f"J{fila}:L{fila}")
            hoja[f"J{fila}"]=info[2][0]+" "+info[2][1]+" "+info[2][2]
            hoja[f"J{fila}"].font= Font("Helvetica", size=10)
            hoja.merge_cells(f"M{fila}:O{fila}")
            hoja[f"M{fila}"]= str(info[3])
            hoja[f"M{fila}"].font= Font("Helvetica", size=10)
            hoja.merge_cells(f"P{fila}:R{fila}")
            hoja[f"P{fila}"]= str(info[4])
            hoja[f"P{fila}"].font= Font("Helvetica", size=10)
            hoja.merge_cells(f"S{fila}:U{fila}")
            hoja[f"S{fila}"] =info[5]
            hoja[f"S{fila}"].font= Font("Helvetica", size=10)
            hoja.merge_cells(f"V{fila}:X{fila}")
            hoja[f"V{fila}"] =str(info[6])
            hoja[f"V{fila}"].font = Font("Helvetica", size=10)
            hoja.merge_cells(f"Y{fila}:AA{fila}")
            hoja[f"Y{fila}"]=str(info[7])
            hoja[f"Y{fila}"].font = Font("Helvetica", size=10)
            hoja.merge_cells(f"AB{fila}:AD{fila}")
            hoja[f"AB{fila}"]= info[8]
            hoja[f"AB{fila}"].font= Font("Helvetica", size=10)
            fila+=1
        archivoBd.save("prueba.xlsx")
        return True
    except:
        return False
def adultoEnlazados():
    """
    Funcionamiento: Crea un reporte con los información que contiene la base datos de los adultos mayores enlazados con voluntarios.
    entrada(dicc): Base de datos con la información de los participantes.
    Salida: Crea o sobreescribe un archivo excel con toda la información de los adultos mayores enlazados con los volutarios.
    """
    documento= Workbook()
    hoja= documento.active
    hoja.merge_cells("A1:L1")
    hoja["A1"].alignment = Alignment(horizontal='center')
    hoja["A1"]="Adultos Enlazados"
    hoja["A1"].font = Font("Helvetica", size=18, bold=True)
    hoja.merge_cells("A2:F2")
    hoja["A2"].alignment=Alignment(horizontal='center')
    hoja["A2"]="Adultos Mayores:"
    hoja["A2"].font = Font("Helvetica", size=15, bold=True)
    hoja.merge_cells("G2:L2")
    hoja["G2"].alignment=Alignment(horizontal='center')
    hoja["G2"]="Voluntarios"
    hoja["G2"].font = Font("Helvetica", size=15, bold=True)
    hoja.merge_cells("A3:C3")
    hoja["A3"].alignment=Alignment(horizontal='center')
    hoja["A3"]="Nombre"
    hoja["A3"].font = Font("Helvetica", size=15, bold=True)
    hoja.merge_cells("D3:F3")
    hoja["D3"].alignment=Alignment(horizontal='center')
    hoja["D3"]= "Código"
    hoja["D3"].font = Font("Helvetica", size=15, bold=True)
    hoja.merge_cells("G3:I3")
    hoja["G3"].alignment=Alignment(horizontal='center')
    hoja["G3"]= "Nombre"
    hoja["G3"].font = Font("Helvetica", size=15, bold=True)
    hoja.merge_cells("J3:L3")
    hoja["J3"].alignment=Alignment(horizontal='center')
    hoja["J3"]= "Código"
    hoja["J3"].font = Font("Helvetica", size=15, bold=True)
    fila=4
    for adulto, voluntario in adultosEnla.items():
        hoja.merge_cells(f"A{fila}:C{fila}")
        hoja[f"A{fila}"]=participantes[adulto][2][0]+" "+participantes[adulto][2][1]+" "+participantes[adulto][2][2]
        hoja[f"A{fila}"].font= Font("Helvetica", size=10)
        hoja.merge(f"D{fila}:F{fila}")
        hoja[f"D{fila}"]=adulto
        hoja[f"D{fila}"].font = Font("Helvetica", size=10)
        hoja.merge_cells(f"G{fila}:I{fila}")
        hoja[f"G{fila}"]= participantes[voluntario][2][0]+" "+participantes[voluntario][2][1]+" "+participantes[voluntario][2][2]
        hoja[f"G{fila}"].font = Font("Helvetica", size=10)
        hoja.merge_cells(f"J{fila}:L{fila}")
        hoja[f"J{fila}"]=voluntario
        hoja[f"J{fila}"].font = Font("Helvetica", size=10)
        fila+=1
    documento.save("reporteEnlazados.xlsx")

